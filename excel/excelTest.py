from calendar import month
from dataclasses import replace
from sre_constants import SUCCESS
import openpyxl
from datetime import datetime, timedelta
import re
from tkinter import messagebox

def replFromat(value,formatString1,formatString2):
    value = value.replace(formatString1,formatString2);
    value = value.replace(" ","");
    return value;


def reSplit(value):
    equipTime = [];
    equipName = [];
    RiChange = [];
    result = [];
    # 설비장비 | 일자 GET

    for i in value:
        i = i.strip();
        
        Ri = i.split("     ")[0]; # 설비명
        Li = i.split("     ")[1]; # 날짜확인

        if "LVAD" in Ri:
            RiChange = replFromat(Ri,"VA","-VA");
            Ri = RiChange;
        elif "JVAD" in Ri:
            RiChange = replFromat(Ri,"VA","-VA");
            Ri = RiChange;
        elif "O" in Ri:
            RiChange = replFromat(Ri,"VA","-VA");
            Ri = RiChange;
        elif "LATH" in Ri:
            RiChange = replFromat(Ri,"LATH","L-Lathe");
            Ri = RiChange;
        elif "LFUR" in Ri:
            RiChange = replFromat(Ri,"LFUR","L-Furnace");
            Ri = RiChange;
        elif "VFUR" in Ri:
            RiChange = replFromat(Ri,"VFUR","V-Furnace");
            Ri = RiChange;
        elif "RFUR" in Ri:
            RiChange = replFromat(Ri,"RFUR","R-Furnace");
            Ri = RiChange;
        if "VLATH" in Ri:
            RiChange = replFromat(Ri,"VLATH","V-Lathe");
            Ri = RiChange;

        equipTime.append(Li);
        equipName.append(Ri);

    for i in zip(equipName, equipTime):
        result.append(i);

    return result;

def todayCell(Today, load_ws,data):
    for cols in load_ws.iter_cols(min_row=6):
        cellAlphabet = str(cols[1]).split(".")[1];
        print(cols[1].value);
        #어제자 제외 및 없음 등 추출 하기 위한
        if cols[1].value == Today:
            sliceToday = cellAlphabet[0:1];
            print(len(data));
            #알파벳 변수로 받고 none / 제외 그대로
            for i in range(7,len(data)+20):
                # 설비명
                eqipNameExcel = load_ws['E'+str(i)].value;
                
                print(eqipNameExcel);
                print("data[j][0] : ", data[j][0]);
                for j in range(7,len(data)):
                    if eqipNameExcel == data[j][0]:
                        # print("시간 데이터 : ", data[j][1]);
                        if data[j][1] == "몇초 전":
                            load_ws[sliceToday+str(i)].value = "";
                        elif "시간 전" in data[j][1]:
                            load_ws[sliceToday+str(i)].value = "";
                        elif "분 전" in data[j][1]:
                            load_ws[sliceToday+str(i)].value = "";
                        elif "하루 전" in data[j][1]:
                            # print("row Check : ", cols[1].value);
                            load_ws[sliceToday+str(i)].value = "";
                        else:
                            load_ws[sliceToday+str(i)].value = data[j][1];

def yesterDayCell(yesterday,load_ws,data,wordBox):
    for cols in load_ws.iter_cols(min_row=6):
        cellAlphabet = str(cols[1]).split(".")[1];
        if cols[1].value == yesterday:
            sliceYesterday =  cellAlphabet[0:1];
            sliceToday = wordBox[0];
            for i in range(7, len(data) + 20):
                # eqipNameExcel = load_ws['E'+str(i)].value;
                yesDt = load_ws[sliceYesterday+str(i)].value
                if load_ws[sliceYesterday+str(i)].value == "제외":
                    #셀 번호 확인
                    cellNumber = str(load_ws[sliceYesterday+str(i)]).split(".")[1].replace(">","")
                    load_ws[sliceToday+str(i)].value = yesDt;
                elif load_ws[sliceYesterday+str(i)].value == "없음":
                    cellNumber = str(load_ws[sliceYesterday+str(i)]).split(".")[1].replace(">","")
                    load_ws[sliceToday+str(i)].value = yesDt;
                elif load_ws[sliceYesterday+str(i)].value == "보류":
                    cellNumber = str(load_ws[sliceYesterday+str(i)]).split(".")[1].replace(">","")
                    load_ws[sliceToday+str(i)].value = yesDt;

def excelReadWrite(path, value):
    data = reSplit(value);
    for i in zip("0", "0"):
        data.insert(0,i);
        data.insert(1,i);
        data.insert(2,i);
        data.insert(3,i);
        data.insert(4,i);
        data.insert(5,i);
        data.insert(6,i);
        data.insert(7,i);
        # data.insert(5,i);
    month = datetime.now().month;
    # 현재 날짜 데이터 관련
    Today = datetime.now();
    yesterday = Today - timedelta(1);

    Today = str(Today).split(" ")[0].replace("-",".");
    yesterday = str(yesterday).split(" ")[0].replace("-",".");
    wordBox = [];
    sheetName = str(month)+'월_일일점검';
    
    print(Today);
    print(yesterday);
    print(sheetName);

    load_monitoring = openpyxl.load_workbook(path);
    load_ws = load_monitoring[sheetName];

    for cols in load_ws.iter_cols(min_row=6):
        cellAlphabet = str(cols[1]).split(".")[1];
        #어제자 제외 및 없음 등 추출 하기 위한
        if cols[1].value == Today:
            sliceToday = cellAlphabet[0:1];
            wordBox.append(sliceToday);

    todayCell(Today,load_ws,data);

    yesterDayCell(yesterday,load_ws,data,wordBox);

    load_monitoring.save(path);

    messagebox.showinfo("확인 창", "excel Input SUCCESS");

        